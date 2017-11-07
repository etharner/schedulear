# -*- coding: utf-8 -*-
from openpyxl import *
import sys, re
import json


def get_merged(sheet, cell):
	global last_merged_range

	idx = cell.coordinate
	for range_ in sheet.merged_cell_ranges:
		merged_cells = list(utils.rows_from_range(range_))
		for row in merged_cells:
			if idx in row:
				if (range_ != last_merged_range):
					last_merged_range = range_
					return sheet.cell(merged_cells[0][0]).value
				else:
					return ''	
        	       
	return sheet[idx].value

def getRow(start_row, row_letter):
	row_data = []
	max_letter = utils.get_column_letter(ws.max_column)
	for row in ws.iter_rows('{}{}:{}{}'.format(row_letter, start_row, max_letter, start_row)):
		for cell in row:
			row_data.append(get_merged(ws, cell))
	return row_data

def getCol(room_row, col_letter):
	days = [u'Понедельник', u'Вторник', u'Среда', u'Четверг', u'Пятница', u'Суббота']
	pairs = ['1\n8:30-10:00', '2\n10:10-11:40', '3\n11:50-13:20', '4\n13:30-15:00', 
		'5\n15:10-16:40', '6\n16:50-18:20', '7\n18:30-20:00']

	col_data = {}
	for d in days:
		col_data[d] = {}
		for p in pairs:
			col_data[d][p] = []

	passed_time = 0
	passed_days = 0
	two_pairs = 0

	for col in range(room_row, ws.max_row - 2):
		if (passed_time > 0):
			data_cell = get_merged(ws, ws[col_letter][col - 1]) 

			if (data_cell != ''):
				col_data[days[passed_days]][pairs[passed_time - 1]].append(data_cell)
				#print(str(col) + ' ' +  days[passed_days] + ' ' + pairs[passed_time - 1][0] + ' ' if data_cell == None 
				#	else str(col) + ' ' +  days[passed_days] + ' ' + pairs[passed_time - 1][0] + ' ' + data_cell)
			two_pairs += 1
			if (two_pairs == 4):
				two_pairs = 0
				passed_time += 1

			if (passed_time > 7):
				passed_time = 0
				passed_days += 1	

		else:
			passed_time = 1

	return col_data	

def parse(room_number):
	global ws, last_merged_range 
	wb = load_workbook('rooms.xlsx')
	sheets = wb.get_sheet_names()
	ws = wb[sheets[0]]

	rooms_row = 10
	rooms_letter = 'D'
	room_row = 12
	room_letter_shift = 4

	last_merged_range = []

	f = open('room.json', 'w')

	row_data = getRow(rooms_row, rooms_letter)
	if (room_number in row_data):
		col_letter = utils.get_column_letter(row_data.index(room_number) + room_letter_shift)
		a = getCol(room_row, col_letter)
		j = json.dumps(a)
		#f.write(j)		
	else:
		a = {}
		j = json.dumps(a)
		#f.write(j)

	return j
